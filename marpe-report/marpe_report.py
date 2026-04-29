"""
MARPE Historical Report
Pulls invoiced MARPE case counts from MagicTouch (Apr 2024 to present),
fits an exponential growth curve, and projects backward to April 2020.
Outputs a branded Excel workbook with three sheets.
"""

import sys
from datetime import date, datetime

import numpy as np
import pandas as pd
import pyodbc
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Connection ──────────────────────────────────────────────────────────────

SERVER = "192.168.10.5"
DATABASE = "DLCPM"
UID = "DLCPM"
PWD = "@chinatown"


def get_connection():
    sql_drivers = [d for d in pyodbc.drivers() if "SQL Server" in d]
    if not sql_drivers:
        raise RuntimeError("No SQL Server ODBC driver found. Install ODBC Driver 17 or 18.")
    driver = sql_drivers[-1]
    cs = f"DRIVER={{{driver}}};SERVER={SERVER};DATABASE={DATABASE};UID={UID};PWD={PWD};TrustServerCertificate=yes;"
    return pyodbc.connect(cs)


# ── Product list ─────────────────────────────────────────────────────────────

MARPE_PRODUCTS = [
    "MARPE LEONE 10",
    "MARPE LEONE 12",
    "MARPE SHIELD SS 12",
    "MARPE SHIELD SS 16",
    "MARPE SHIELD TS 10",
    "MARPE SHIELD TS 12",
    "MARPE SHIELD TS 14",
    "MARPE SHIELD TS 16",
    "MARPE SS 12",
    "MARPE SS 16",
    "MARPE SWAP TS 12",
    "MARPE SWAP TS 16",
    "MARPE TIGER 8",
    "MARPE TIGER 10",
    "MARPE TIGER 12",
    "MARPE TIGER 14",
    "MARPE TIGER 16",
    "MSE 8mm",
    "MSE 10mm",
    "MSE 12mm",
    "Ti MARPE",
]

FAMILIES = {
    "TIGER":  ["MARPE TIGER 8", "MARPE TIGER 10", "MARPE TIGER 12", "MARPE TIGER 14", "MARPE TIGER 16"],
    "LEONE":  ["MARPE LEONE 10", "MARPE LEONE 12"],
    "SHIELD": ["MARPE SHIELD SS 12", "MARPE SHIELD SS 16",
               "MARPE SHIELD TS 10", "MARPE SHIELD TS 12", "MARPE SHIELD TS 14", "MARPE SHIELD TS 16"],
    "SS":     ["MARPE SS 12", "MARPE SS 16"],
    "SWAP":   ["MARPE SWAP TS 12", "MARPE SWAP TS 16"],
    "MSE":    ["MSE 8mm", "MSE 10mm", "MSE 12mm"],
    "OTHER":  ["Ti MARPE"],
}

# ── Brand colors ─────────────────────────────────────────────────────────────

RED       = "C41227"
WHITE     = "FFFFFF"
BLACK     = "000000"
SMOKE     = "F5F5F5"
LGRAY     = "E8E8E8"
DGRAY     = "888888"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, italic=False, color=BLACK, size=11):
    return Font(bold=bold, italic=italic, color=color, size=size)

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(bottom=s)

# ── SQL query ─────────────────────────────────────────────────────────────────

SQL = """
SELECT
    FORMAT(ca.InvoiceDate, 'yyyy-MM') AS InvoiceMonth,
    cp.ProductID,
    COUNT(*) AS Cnt
FROM dbo.caseproducts cp
JOIN dbo.cases ca ON cp.CaseID = ca.CaseID
WHERE cp.ProductID IN ({placeholders})
  AND ca.Status = 'Invoiced'
  AND ca.Deleted = 0
  AND ca.InvoiceDate IS NOT NULL
GROUP BY FORMAT(ca.InvoiceDate, 'yyyy-MM'), cp.ProductID
ORDER BY InvoiceMonth, cp.ProductID
"""


def fetch_actuals(conn) -> pd.DataFrame:
    placeholders = ", ".join("?" * len(MARPE_PRODUCTS))
    query = SQL.format(placeholders=placeholders)
    cursor = conn.cursor()
    cursor.execute(query, MARPE_PRODUCTS)
    rows = cursor.fetchall()
    cols = [d[0] for d in cursor.description]
    return pd.DataFrame.from_records(rows, columns=cols)


# ── Pivot + regression ────────────────────────────────────────────────────────

def build_actual_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(index="InvoiceMonth", columns="ProductID", values="Cnt", fill_value=0)
    pivot = pivot.reindex(columns=MARPE_PRODUCTS, fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot.index = pd.PeriodIndex(pivot.index, freq="M")
    return pivot


def fit_and_project(actual: pd.DataFrame) -> pd.DataFrame:
    """
    Fit log-linear regression on complete months (exclude current partial month).
    Project back to April 2020.
    Distribute projected totals using the product mix from the first 3 actual months.
    """
    today = date.today()
    current_period = pd.Period(today, freq="M")

    # Complete months only for fitting
    complete = actual[actual.index < current_period].copy()
    totals = complete["Total"].values.astype(float)

    # Exclude any zero months that would break log
    nonzero_mask = totals > 0
    x = np.arange(len(totals))[nonzero_mask]
    y = totals[nonzero_mask]
    log_y = np.log(y)
    slope, intercept = np.polyfit(x, log_y, 1)

    # Project back to 2020-04
    start_actual = actual.index[0]  # e.g. 2024-04
    start_proj = pd.Period("2020-04", freq="M")
    n_back = (start_actual - start_proj).n  # number of months to project back

    proj_periods = pd.period_range(start=start_proj, periods=n_back, freq="M")
    proj_x = np.arange(-n_back, 0)  # negative offsets relative to first actual month
    proj_totals = np.exp(slope * proj_x + intercept)
    proj_totals = np.maximum(proj_totals, 1).round().astype(int)

    # Product mix from first 3 actual months
    n_seed = min(3, len(complete))
    seed = complete.iloc[:n_seed][MARPE_PRODUCTS].sum()
    seed_sum = seed.sum()
    if seed_sum == 0:
        proportions = pd.Series(1 / len(MARPE_PRODUCTS), index=MARPE_PRODUCTS)
    else:
        proportions = seed / seed_sum

    rows = []
    for period, total in zip(proj_periods, proj_totals):
        # Allow 0 for near-zero-proportion products (max(1,...) caused negatives on small totals)
        row = {p: max(0, round(proportions[p] * total)) for p in MARPE_PRODUCTS}
        diff = total - sum(row.values())
        # Spread rounding correction across highest-count products to avoid any going negative
        if diff > 0:
            for p in sorted(MARPE_PRODUCTS, key=lambda p: proportions[p], reverse=True):
                if diff == 0:
                    break
                row[p] += 1
                diff -= 1
        elif diff < 0:
            for p in sorted(MARPE_PRODUCTS, key=lambda p: row[p], reverse=True):
                if diff == 0:
                    break
                take = min(-diff, row[p])
                row[p] -= take
                diff += take
        row["Total"] = total
        rows.append(row)

    proj_df = pd.DataFrame(rows, index=proj_periods)
    return proj_df


def add_family_cols(df: pd.DataFrame) -> pd.DataFrame:
    fam_df = pd.DataFrame(index=df.index)
    for fam, products in FAMILIES.items():
        cols = [p for p in products if p in df.columns]
        fam_df[fam] = df[cols].sum(axis=1)
    fam_df["Total"] = fam_df.sum(axis=1)
    return fam_df


# ── Excel helpers ─────────────────────────────────────────────────────────────

def write_header_row(ws, headers, row=1):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = fill(RED)
        cell.font = font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def write_data_rows(ws, data_df, projected_index_set, start_row=2):
    for r_idx, (period, row) in enumerate(data_df.iterrows(), start=start_row):
        is_proj = str(period) in projected_index_set
        row_fill = fill(SMOKE) if is_proj else fill(WHITE)
        month_label = str(period)
        if is_proj:
            month_label += " (Est.)"

        for c_idx, val in enumerate([month_label] + list(row.values), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = row_fill
            # Month column
            if c_idx == 1:
                cell.font = font(italic=is_proj, color=DGRAY if is_proj else BLACK)
                cell.alignment = Alignment(horizontal="left")
            # Total column (last)
            elif c_idx == len(row) + 1:
                cell.fill = fill(LGRAY)
                cell.font = font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = font(italic=is_proj, color=DGRAY if is_proj else BLACK)
                cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

    return r_idx  # last row written


def auto_fit_columns(ws, min_width=8, max_width=20):
    for col in ws.columns:
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_width, max(min_width, max(lengths) + 2)
            )


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_by_product_sheet(wb, full_df, projected_periods):
    ws = wb.create_sheet("By Product")
    ws.sheet_properties.tabColor = RED
    headers = ["Month"] + MARPE_PRODUCTS + ["Total"]
    write_header_row(ws, headers)
    write_data_rows(ws, full_df, projected_periods)
    ws.freeze_panes = "B2"
    auto_fit_columns(ws)


def write_by_family_sheet(wb, proj_family, actual_family, projected_periods):
    ws = wb.create_sheet("By Family")
    ws.sheet_properties.tabColor = LGRAY
    fam_names = list(FAMILIES.keys())
    headers = ["Month"] + fam_names + ["Total"]
    write_header_row(ws, headers)
    full_fam = pd.concat([proj_family, actual_family])
    write_data_rows(ws, full_fam, projected_periods)
    ws.freeze_panes = "B2"
    auto_fit_columns(ws)


def write_summary_sheet(wb, actual: pd.DataFrame, proj: pd.DataFrame, slope: float, intercept: float):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = BLACK

    # Title banner
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "MARPE Historical Report — Partners Dental Solutions"
    title_cell.fill = fill(RED)
    title_cell.font = font(bold=True, color=WHITE, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Generated {datetime.today().strftime('%B %d, %Y')}"
    sub.fill = fill(SMOKE)
    sub.font = font(italic=True, color=DGRAY)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    def label_val(row, label, value, bold_val=True):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = font(bold=True)
        lc.fill = fill(SMOKE)
        lc.alignment = Alignment(horizontal="left")
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = font(bold=bold_val, size=12 if bold_val else 11)
        vc.fill = fill(WHITE)
        vc.alignment = Alignment(horizontal="left")

    actual_total = int(actual["Total"].sum())
    proj_total   = int(proj["Total"].sum())
    grand_total  = actual_total + proj_total

    today = date.today()
    current_period = pd.Period(today, freq="M")
    complete_actual = actual[actual.index < current_period]
    monthly_avg = complete_actual["Total"].mean()
    monthly_min = complete_actual["Total"].min()
    monthly_max = complete_actual["Total"].max()

    r = 4
    label_val(r,     "Data start (actual)",         str(actual.index[0]))
    label_val(r + 1, "Data end (actual)",            str(actual.index[-1]))
    label_val(r + 2, "Total actual (in database)",   f"{actual_total:,}")
    label_val(r + 3, "Projection start",             "April 2020")
    label_val(r + 4, "Projected months",             f"{len(proj)}")
    label_val(r + 5, "Total projected (2020–2024)",  f"{proj_total:,}")
    label_val(r + 6, "GRAND TOTAL ESTIMATE",         f"{grand_total:,}")
    label_val(r + 7, "Monthly avg (actual)",         f"{monthly_avg:.0f}")
    label_val(r + 8, "Monthly range (actual)",       f"{monthly_min:,} – {monthly_max:,}")

    ws.cell(r + 6, 1).fill = fill(RED)
    ws.cell(r + 6, 1).font = font(bold=True, color=WHITE)
    ws.cell(r + 6, 2).fill = fill(RED)
    ws.cell(r + 6, 2).font = font(bold=True, color=WHITE, size=14)

    # Methodology note
    note_row = r + 11
    ws.merge_cells(f"A{note_row}:D{note_row}")
    hdr = ws.cell(note_row, 1, "Projection Methodology")
    hdr.fill = fill(LGRAY)
    hdr.font = font(bold=True)

    monthly_rate_pct = (np.exp(slope) - 1) * 100
    note = (
        f"Actual monthly counts from {actual.index[0]} to {actual.index[-1]} were used to fit a "
        f"log-linear (exponential) regression model. The model estimates a monthly growth rate of "
        f"{monthly_rate_pct:.1f}%. This model was then applied backward 48 months to April 2020 "
        f"to estimate pre-database volumes. Product mix for projected months is derived from the "
        f"product proportions observed in the first 3 months of actual data. "
        f"Projected figures carry significant uncertainty — actual historical volume may differ."
    )
    ws.merge_cells(f"A{note_row+1}:D{note_row+4}")
    nc = ws.cell(note_row + 1, 1, note)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    nc.font = font(italic=True, color=DGRAY, size=10)
    ws.row_dimensions[note_row + 1].height = 80

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12


# ── Summary v2 ───────────────────────────────────────────────────────────────

def write_summary_v2_sheet(wb, actual: pd.DataFrame, proj: pd.DataFrame, slope: float, intercept: float):
    ws = wb.create_sheet("Summary v2")
    ws.sheet_properties.tabColor = RED

    today = date.today()
    current_period = pd.Period(today, freq="M")
    complete_actual = actual[actual.index < current_period]

    actual_total  = int(actual["Total"].sum())
    proj_total    = int(proj["Total"].sum())
    grand_total   = actual_total + proj_total
    monthly_avg   = complete_actual["Total"].mean()
    monthly_min   = int(complete_actual["Total"].min())
    monthly_max   = int(complete_actual["Total"].max())
    min_month     = str(complete_actual["Total"].idxmin())
    max_month     = str(complete_actual["Total"].idxmax())
    monthly_rate  = (np.exp(slope) - 1) * 100
    annual_rate   = (np.exp(slope * 12) - 1) * 100
    doubling_mo   = round(np.log(2) / slope)

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28

    row = [0]  # mutable counter

    def next_row(n=1):
        row[0] += n
        return row[0]

    def r():
        return row[0]

    def section_header(title, col_span=5):
        rr = next_row()
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=col_span)
        c = ws.cell(rr, 1, title)
        c.fill = fill(RED)
        c.font = font(bold=True, color=WHITE, size=12)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[rr].height = 22
        return rr

    def sub_header(title, col_span=5):
        rr = next_row()
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=col_span)
        c = ws.cell(rr, 1, title)
        c.fill = fill(LGRAY)
        c.font = font(bold=True, color=BLACK, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[rr].height = 18
        return rr

    def kv(label, value, note="", bold_val=True, highlight=False):
        rr = next_row()
        lc = ws.cell(rr, 1, label)
        lc.font = font(bold=True, size=10)
        lc.fill = fill(SMOKE)
        lc.alignment = Alignment(horizontal="left", indent=1)
        vc = ws.cell(rr, 2, value)
        vc.font = font(bold=bold_val, size=11 if bold_val else 10)
        vc.fill = fill(RED if highlight else WHITE)
        vc.font = Font(bold=bold_val, color=WHITE if highlight else BLACK, size=13 if highlight else (11 if bold_val else 10))
        vc.alignment = Alignment(horizontal="left")
        if note:
            nc = ws.cell(rr, 3, note)
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=5)
            nc.font = font(italic=True, color=DGRAY, size=9)
            nc.alignment = Alignment(horizontal="left", vertical="center")

    def spacer():
        next_row()

    def table_header_row(*cols):
        rr = next_row()
        for ci, col in enumerate(cols, 1):
            c = ws.cell(rr, ci, col)
            c.fill = fill(BLACK)
            c.font = font(bold=True, color=WHITE, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[rr].height = 18

    def table_data_row(*vals, shaded=False):
        rr = next_row()
        bg = fill(SMOKE) if shaded else fill(WHITE)
        for ci, v in enumerate(vals, 1):
            c = ws.cell(rr, ci, v)
            c.fill = bg
            c.font = font(size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            c.border = thin_border()

    # ── Title banner ──
    rr = next_row()
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    tc = ws.cell(rr, 1, "MARPE Historical Report — Partners Dental Solutions")
    tc.fill = fill(RED)
    tc.font = font(bold=True, color=WHITE, size=16)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rr].height = 40

    rr = next_row()
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    sc = ws.cell(rr, 1, f"Generated {datetime.today().strftime('%B %d, %Y')}  |  Data: {actual.index[0]} to {actual.index[-1]}  |  Projection: April 2020 to {proj.index[-1]}")
    sc.fill = fill(SMOKE)
    sc.font = font(italic=True, color=DGRAY, size=10)
    sc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[rr].height = 18

    spacer()

    # ── Section 1: The Bottom Line ──
    section_header("THE BOTTOM LINE")
    kv("Grand Total Estimate (Apr 2020 - today)", f"{grand_total:,}", "confirmed + projected combined", highlight=True)
    kv("  Confirmed in database (invoiced cases)", f"{actual_total:,}", f"Apr 2024 to {actual.index[-1]} — exact count from MagicTouch")
    kv("  Estimated / projected (pre-database)", f"{proj_total:,}", "Apr 2020 to Mar 2024 — modeled from growth trend")
    kv("  Confidence in confirmed number", "High", "directly queried from MagicTouch, Status = Invoiced, not deleted")
    kv("  Confidence in projected number", "Moderate", "exponential model; uncertainty grows further back in time")

    spacer()

    # ── Section 2: Year-by-Year Breakdown ──
    section_header("YEAR-BY-YEAR BREAKDOWN")

    # Build year rows from proj + actual
    full = pd.concat([proj, actual])
    full_by_year = full.groupby(full.index.year)["Total"].sum()

    year_labels = {
        2020: ("Apr–Dec 2020", "Projected", "9 months — estimated ramp-up from launch"),
        2021: ("Full Year 2021", "Projected", "First full year of MARPE production"),
        2022: ("Full Year 2022", "Projected", "Growth period"),
        2023: ("Full Year 2023", "Projected", "Growth period"),
        2024: ("Full Year 2024", "Mixed", "Jan–Mar projected; Apr–Dec in database"),
        2025: ("Full Year 2025", "Actual", "Full year in MagicTouch database"),
        2026: (f"Jan–{today.strftime('%b')} 2026", "Actual (partial)", f"Partial year, through {today.strftime('%B %d, %Y')}"),
    }

    table_header_row("Year / Period", "Est. Count", "Status", "% of Total", "Notes")
    for yr, total in full_by_year.items():
        label, status, note = year_labels.get(yr, (str(yr), "Actual", ""))
        pct = total / grand_total * 100
        shaded = (yr % 2 == 0)
        table_data_row(label, f"{int(total):,}", status, f"{pct:.1f}%", note, shaded=shaded)

    spacer()

    # ── Section 3: Actual Data Highlights ──
    section_header("ACTUAL DATA HIGHLIGHTS  (MagicTouch database)")
    kv("Database coverage", f"{actual.index[0]} to {actual.index[-1]}", f"{len(actual)} months total")
    kv("Complete months used for modeling", str(len(complete_actual)), "current partial month excluded from regression")
    kv("Busiest month on record", f"{monthly_max:,} units", f"{max_month}")
    kv("Slowest month on record", f"{monthly_min:,} units", f"{min_month}")
    kv("Average per month (actual)", f"{monthly_avg:.0f} units", "across all complete months")
    kv("Implied annual run rate (current avg)", f"{monthly_avg * 12:,.0f} units/year", "based on recent monthly average")

    # Year totals for actual years
    spacer()
    sub_header("Annual Totals from Database")
    table_header_row("Year", "Total Units", "Months of Data", "Monthly Avg", "Notes")
    actual_by_year = complete_actual.groupby(complete_actual.index.year)["Total"].agg(["sum", "count", "mean"])
    for yr, row_data in actual_by_year.iterrows():
        shaded = (yr % 2 == 0)
        note = "partial year (Apr-Dec)" if yr == 2024 else ("partial year (Jan-Apr)" if yr == today.year else "full year")
        table_data_row(str(yr), f"{int(row_data['sum']):,}", str(int(row_data['count'])), f"{row_data['mean']:.0f}", note, shaded=shaded)

    spacer()

    # ── Section 4: Growth Trend ──
    section_header("GROWTH TREND ANALYSIS")
    kv("Monthly growth rate (modeled)", f"{monthly_rate:.1f}%", "derived from exponential regression on Apr 2024 - present")
    kv("Implied annual growth rate", f"{annual_rate:.0f}%", "compounded monthly rate over 12 months")
    kv("Implied volume doubling time", f"~{doubling_mo} months", "at current growth rate")
    kv("2024 database total (Apr-Dec)", f"{int(complete_actual[complete_actual.index.year == 2024]['Total'].sum()):,}", "9 months of actual data")
    kv("2025 full-year total", f"{int(complete_actual[complete_actual.index.year == 2025]['Total'].sum()):,}", "12 months of actual data")
    yr25 = complete_actual[complete_actual.index.year == 2025]["Total"].sum()
    yr24 = complete_actual[complete_actual.index.year == 2024]["Total"].sum()
    if yr24 > 0:
        yoy = (yr25 / yr24 - 1) * 100
        kv("2024-to-2025 volume growth", f"+{yoy:.0f}%", "note: 2024 is only 9 months vs 2025 full year")

    spacer()

    # ── Section 5: Top Products (actual data) ──
    section_header("TOP PRODUCTS  (actual data only, Apr 2024 - present)")
    product_totals = actual[MARPE_PRODUCTS].sum().sort_values(ascending=False)
    table_header_row("Product", "Total Units", "% of Actuals", "Avg / Month", "Family")
    family_lookup = {p: fam for fam, prods in FAMILIES.items() for p in prods}
    for i, (prod, total) in enumerate(product_totals.items()):
        if total == 0:
            continue
        pct = total / actual_total * 100
        avg = total / len(complete_actual)
        fam = family_lookup.get(prod, "OTHER")
        table_data_row(prod, f"{int(total):,}", f"{pct:.1f}%", f"{avg:.1f}", fam, shaded=(i % 2 == 0))

    spacer()

    # ── Section 6: Methodology ──
    section_header("HOW THE PROJECTION WORKS")

    steps = [
        ("Step 1 — Query the database",
         "All cases in dbo.cases joined to dbo.caseproducts where Status = 'Invoiced' and Deleted = 0. "
         "Only the 21 MARPE/MSE product codes are included. Results are grouped by invoice month."),
        ("Step 2 — Build a monthly time series",
         f"The query returns {len(actual)} months of data from {actual.index[0]} to {actual.index[-1]}. "
         "Each month has a count of how many MARPE units were invoiced across all products."),
        ("Step 3 — Fit an exponential growth model",
         f"We take the natural log of each month's total and fit a straight line (linear regression) "
         f"through those log values. This is equivalent to fitting an exponential curve to the raw counts. "
         f"The result: a monthly growth rate of {monthly_rate:.1f}%, or ~{annual_rate:.0f}% per year."),
        ("Step 4 — Project backward to April 2020",
         f"The fitted model is run backward 48 months to April 2020 — when MARPE production began. "
         f"At that point the model estimates approximately {int(np.exp(slope * -48 + intercept)):,} units/month. "
         "This is plausible for a product in its first month of production."),
        ("Step 5 — Distribute projected totals across products",
         "The projected monthly total is split across individual products using the product mix "
         "observed in the first 3 months of real data (Apr–Jun 2024). Products that were more popular "
         "early in the data get a proportionally larger share of projected months."),
        ("Step 6 — Caveats and uncertainty",
         "Exponential models assume a constant percentage growth rate throughout the entire period. "
         "In reality, early-stage products often have slower initial ramp-ups followed by faster growth. "
         "This means the projection likely underestimates 2022-2023 and overestimates 2020-2021. "
         "The grand total estimate should be treated as a reasonable order-of-magnitude figure, not a precise count."),
    ]

    for step_title, step_body in steps:
        sub_header(step_title)
        rr = next_row(2)
        ws.merge_cells(start_row=rr - 1, start_column=1, end_row=rr, end_column=5)
        c = ws.cell(rr - 1, 1, step_body)
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        c.font = font(size=10, color=BLACK)
        ws.row_dimensions[rr - 1].height = 45
        ws.row_dimensions[rr].height = 5


# ── Charts ───────────────────────────────────────────────────────────────────

def write_charts_sheet(wb, n_proj, n_actual):
    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = RED

    fam_ws = wb["By Family"]

    # Row positions in By Family: row 1 = header, rows 2..n_proj+1 = projected, then actual
    hdr_row   = 1
    data_start = 2
    data_end  = n_proj + n_actual + 1   # last data row
    n_fam     = len(FAMILIES)           # 7 family columns (B-H)
    total_col = n_fam + 2               # column I = 9

    # PDS palette for family stacks (TIGER→OTHER)
    family_palette = ["C41227", "1A1A1A", "555555", "888888", "AAAAAA", "CCCCCC", "E0E0E0"]

    # ── Banner ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    banner = ws["A1"]
    banner.value = "MARPE Production Charts — Partners Dental Solutions"
    banner.fill = fill(RED)
    banner.font = font(bold=True, color=WHITE, size=14)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:P2")
    sub = ws["A2"]
    sub.value = (
        f"Data: {fam_ws.cell(n_proj + 2, 1).value} to {fam_ws.cell(data_end, 1).value} (actual) | "
        f"Projected: {fam_ws.cell(2, 1).value} to {fam_ws.cell(n_proj + 1, 1).value}"
    )
    sub.fill = fill(SMOKE)
    sub.font = font(italic=True, color=DGRAY, size=10)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Chart 1: Monthly Total — Area Chart ─────────────────────────────────
    chart1 = AreaChart()
    chart1.grouping = "standard"
    chart1.title = "Monthly MARPE Units — Apr 2020 to Present"
    chart1.style = 2
    chart1.y_axis.title = "Units / Month"
    chart1.x_axis.title = "Month"
    chart1.height = 14
    chart1.width = 28

    # Categories: Month column, all data rows
    cats1 = Reference(fam_ws, min_col=1, min_row=data_start, max_row=data_end)
    # Total series (row 1 = header used as series name)
    data1 = Reference(fam_ws, min_col=total_col, min_row=hdr_row, max_row=data_end)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)

    s1 = chart1.series[0]
    s1.graphicalProperties.solidFill = "F5A8B0"       # light red fill under curve
    s1.graphicalProperties.line.solidFill = RED        # PDS red line on top
    s1.graphicalProperties.line.width = 20000          # ~1.6 pt

    # Show every 6th x-axis label so it's readable (73 months is dense)
    chart1.x_axis.tickLblSkip = 6
    chart1.x_axis.tickMarkSkip = 1

    ws.add_chart(chart1, "A4")

    # Section label below chart 1
    lbl1_row = 33
    ws.merge_cells(f"A{lbl1_row}:P{lbl1_row}")
    lbl1 = ws.cell(lbl1_row, 1, "Note: Months before April 2024 are estimated via exponential projection — see Summary v2 for methodology.")
    lbl1.fill = fill(SMOKE)
    lbl1.font = font(italic=True, color=DGRAY, size=9)
    lbl1.alignment = Alignment(horizontal="center")

    # ── Chart 2: Family Breakdown — Stacked Bar ──────────────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "MARPE by Product Family — Full History (Apr 2020 to Present)"
    chart2.style = 2
    chart2.y_axis.title = "Units"
    chart2.x_axis.title = "Month"
    chart2.height = 14
    chart2.width = 28

    cats2 = Reference(fam_ws, min_col=1, min_row=data_start, max_row=data_end)
    # Family columns B-H, include header row for series names
    data2 = Reference(fam_ws, min_col=2, min_row=hdr_row, max_row=data_end, max_col=n_fam + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    for s, color in zip(chart2.series, family_palette):
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color

    chart2.x_axis.tickLblSkip = 6
    chart2.x_axis.tickMarkSkip = 1

    ws.add_chart(chart2, "A35")

    ws.merge_cells(f"A64:P64")
    lbl2 = ws.cell(64, 1, "TIGER family (red) dominates volume. First 48 months are projected using early product-mix proportions.")
    lbl2.fill = fill(SMOKE)
    lbl2.font = font(italic=True, color=DGRAY, size=9)
    lbl2.alignment = Alignment(horizontal="center")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to database…")
    conn = get_connection()

    print("Fetching MARPE actuals…")
    df_raw = fetch_actuals(conn)
    conn.close()

    if df_raw.empty:
        print("No data returned. Check product names and filters.")
        sys.exit(1)

    print(f"  Rows fetched: {len(df_raw)}")
    actual = build_actual_pivot(df_raw)
    print(f"  Months in DB: {len(actual)}  ({actual.index[0]} to {actual.index[-1]})")
    print(f"  Total invoiced MARPEs: {actual['Total'].sum():,}")

    print("Fitting exponential regression and projecting backward…")
    today = date.today()
    current_period = pd.Period(today, freq="M")
    complete = actual[actual.index < current_period]
    totals = complete["Total"].values.astype(float)
    nonzero_mask = totals > 0
    x = np.arange(len(totals))[nonzero_mask]
    y = totals[nonzero_mask]
    slope, intercept = np.polyfit(x, np.log(y), 1)

    proj = fit_and_project(actual)
    print(f"  Projected months: {len(proj)}  ({proj.index[0]} to {proj.index[-1]})")
    print(f"  Total projected: {proj['Total'].sum():,}")
    print(f"  Grand total estimate: {actual['Total'].sum() + proj['Total'].sum():,}")

    projected_periods = {str(p) for p in proj.index}
    full_df = pd.concat([proj[MARPE_PRODUCTS + ["Total"]], actual[MARPE_PRODUCTS + ["Total"]]])

    actual_family = add_family_cols(actual)
    proj_family   = add_family_cols(proj)

    print("Building Excel workbook…")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_by_product_sheet(wb, full_df, projected_periods)
    write_by_family_sheet(wb, proj_family, actual_family, projected_periods)
    write_summary_sheet(wb, actual, proj, slope, intercept)
    write_summary_v2_sheet(wb, actual, proj, slope, intercept)
    write_charts_sheet(wb, len(proj), len(actual))

    base = f"marpe_report_{date.today().isoformat()}"
    out_path = f"{base}.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = f"{base}_v2.xlsx"
        wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
